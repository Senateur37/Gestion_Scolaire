from django.shortcuts import render
from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse
from students.models import Student
from teachers.models import Teacher
from payments.models import Payment
from academics.models import SchoolClass
from django.db.models import Count, Sum
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io

class AdminRequiredMixin(UserPassesTestMixin):
    def test_func(self):
        return self.request.user.role == 'ADMIN'

class AnalyticsHubView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request):
        # 1. Demographics
        total_students = Student.objects.count()
        total_teachers = Teacher.objects.count()
        demographics_data = [total_students, total_teachers]
        
        # 2. Finance
        total_expected = Payment.objects.aggregate(Sum('amount'))['amount__sum'] or 0
        total_collected = Payment.objects.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
        total_remaining = float(total_expected) - float(total_collected)
        
        finance_data = [float(total_collected), float(total_remaining)]
        
        # 3. Academics (Students per class)
        classes = SchoolClass.objects.annotate(student_count=Count('enrollment'))
        class_names = [c.name for c in classes]
        class_counts = [c.student_count for c in classes]

        # 4. Human Resources
        from hr.models import Payroll, LeaveRequest
        paid_payroll = Payroll.objects.filter(is_paid=True).aggregate(Sum('net_salary'))['net_salary__sum'] or 0
        unpaid_payroll = Payroll.objects.filter(is_paid=False).aggregate(Sum('net_salary'))['net_salary__sum'] or 0
        hr_payroll_data = [float(paid_payroll), float(unpaid_payroll)]

        leave_approved = LeaveRequest.objects.filter(status='APPROVED').count()
        leave_pending = LeaveRequest.objects.filter(status='PENDING').count()
        leave_rejected = LeaveRequest.objects.filter(status='REJECTED').count()
        hr_leave_data = [leave_approved, leave_pending, leave_rejected]

        import json

        context = {
            'demographics_data': json.dumps(demographics_data),
            'finance_data': json.dumps(finance_data),
            'class_names': json.dumps(class_names),
            'class_counts': json.dumps(class_counts),
            'hr_payroll_data': json.dumps(hr_payroll_data),
            'hr_leave_data': json.dumps(hr_leave_data),
        }
        return render(request, 'reports/analytics_hub.html', context)

class ExportExcelView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="rapport_financier.xlsx"'

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Paiements'

        # Headers
        columns = ['ID', 'Élève', 'Montant Total (€)', 'Montant Versé (€)', 'Reste à Payer (€)', 'Date d\'échéance', 'Dernier paiement', 'Statut']
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column_title

        # Data
        payments = Payment.objects.all().select_related('student__user')
        for row_num, payment in enumerate(payments, 2):
            worksheet.cell(row=row_num, column=1, value=payment.id)
            worksheet.cell(row=row_num, column=2, value=payment.student.user.get_full_name())
            worksheet.cell(row=row_num, column=3, value=float(payment.amount))
            worksheet.cell(row=row_num, column=4, value=float(payment.amount_paid))
            worksheet.cell(row=row_num, column=5, value=float(payment.remaining_amount))
            worksheet.cell(row=row_num, column=6, value=payment.due_date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row_num, column=7, value=payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '-')
            worksheet.cell(row=row_num, column=8, value=payment.get_status_display())

        workbook.save(response)
        return response

class ExportPDFView(LoginRequiredMixin, AdminRequiredMixin, View):
    def get(self, request):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="liste_eleves.pdf"'

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        elements.append(Paragraph("Liste Complète des Élèves", styles['Title']))
        elements.append(Spacer(1, 20))

        data = [['ID', 'Nom Complet', 'Sexe', 'Date de Naissance']]
        for student in Student.objects.all().select_related('user'):
            data.append([
                str(student.user.id),
                student.user.get_full_name(),
                student.gender,
                student.birth_date.strftime('%d/%m/%Y') if getattr(student, 'birth_date', None) else '-'
            ])

        table = Table(data, colWidths=[50, 200, 100, 150])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1'))
        ]))
        
        elements.append(table)
        doc.build(elements)

        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)

        return response
